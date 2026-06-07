import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import database
from components.custom_styles import inject_custom_css
from components.export_utils import export_to_csv, export_to_excel, export_to_pdf

def io_export_logs_excel(df: pd.DataFrame) -> bytes:
    """Helper to export audit logs to Excel."""
    import io
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Export a clean display format
        cols = ["Log ID", "Action", "Task ID", "Task Title", "Executed By", "Timestamp", "Details"]
        df_exp = df[cols].copy()
        df_exp["Timestamp"] = df_exp["Timestamp"].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if isinstance(x, datetime) else str(x))
        
        df_exp.to_excel(writer, sheet_name='Audit Logs', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Audit Logs']
        
        # Styles
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid') # Dark Slate
        cell_font = Font(name='Segoe UI', size=10)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        for col_idx in range(1, len(cols) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        for row_idx in range(2, len(df_exp) + 2):
            for col_idx in range(1, len(cols) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                if col_idx in [1, 2, 3, 6]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    return output.getvalue()

def io_export_logs_pdf(df: pd.DataFrame) -> bytes:
    """Helper to export audit logs to PDF using reportlab."""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from config import FOOTER_CREATOR, APP_NAME, TAGLINE
    
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=letter,
        rightMargin=40, leftMargin=40, 
        topMargin=54, bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#475569'),
        spaceAfter=15
    )
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=8,
        leading=9,
        textColor=colors.HexColor('#1E293B')
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white
    )
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#94A3B8'),
        alignment=1
    )
    
    story = []
    story.append(Paragraph(f"{APP_NAME} - Audit Log Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Filters applied", subtitle_style))
    story.append(Spacer(1, 10))
    
    table_headers = ["Action", "Task ID", "Task Title", "User", "Timestamp", "Details"]
    table_data = [[Paragraph(h, header_style) for h in table_headers]]
    
    for _, row in df.iterrows():
        action = str(row['Action'])
        task_id = str(row['Task ID'])
        title_text = str(row['Task Title'])
        if len(title_text) > 30:
            title_text = title_text[:27] + "..."
        user = str(row['Executed By'])
        timestamp = row['Timestamp'].strftime('%m-%d %H:%M') if isinstance(row['Timestamp'], datetime) else str(row['Timestamp'])
        details = str(row['Details'])
        if len(details) > 40:
            details = details[:37] + "..."
            
        row_cells = [
            Paragraph(action, body_style),
            Paragraph(task_id, body_style),
            Paragraph(title_text, body_style),
            Paragraph(user, body_style),
            Paragraph(timestamp, body_style),
            Paragraph(details, body_style)
        ]
        table_data.append(row_cells)
        
    # Printable area is 8.5" width - 80 margin = 7.5"
    task_table = Table(table_data, colWidths=[0.8*inch, 0.6*inch, 1.8*inch, 1.0*inch, 1.1*inch, 2.2*inch])
    task_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(task_table)
    story.append(Spacer(1, 30))
    story.append(Paragraph(f"{FOOTER_CREATOR} | TaskTracker Pro &copy; 2026", footer_style))
    
    doc.build(story)
    return pdf_buffer.getvalue()

inject_custom_css()

st.title("🔒 Security Audit Trail")
st.write("Permanent, immutable log of all task activities and user modifications.")

# Fetch audit logs
logs = database.get_audit_logs()

if not logs:
    st.info("No logs registered in the database yet. Perform task operations to generate logs.")
else:
    # Build dataframe
    log_data = []
    for log in logs:
        log_data.append({
            "Log ID": log.log_id,
            "Action": log.action_type,
            "Task ID": log.task_id,
            "Task Title": log.task_title,
            "Executed By": log.user_name,
            "Timestamp": log.timestamp,
            "Date": log.timestamp.date() if log.timestamp else None,
            "Details": log.details or ""
        })
    df_logs = pd.DataFrame(log_data)
    
    st.markdown("<div class='section-header'>Filter Audit Logs</div>", unsafe_allow_html=True)
    
    # Filter Controls
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        search_query = st.text_input("🔍 Search (Task Title or Details)", placeholder="Enter keyword...")
    with col_f2:
        # Date range filter
        min_date = df_logs["Date"].min() if not df_logs.empty else date.today()
        max_date = df_logs["Date"].max() if not df_logs.empty else date.today()
        
        # Default date range
        date_range = st.date_input(
            "Select Date Range", 
            value=(min_date, max_date),
            min_value=min_date - timedelta(days=365),
            max_value=max_date + timedelta(days=30)
        )
        
    col_f3, col_f4 = st.columns(2)
    with col_f3:
        all_users = sorted(df_logs["Executed By"].dropna().unique())
        selected_users = st.multiselect("Filter by User Name", options=all_users, default=all_users)
    with col_f4:
        all_actions = sorted(df_logs["Action"].dropna().unique())
        selected_actions = st.multiselect("Filter by Action Type", options=all_actions, default=all_actions)
        
    # Apply filters
    filtered_df = df_logs.copy()
    
    # 1. Search Query
    if search_query:
        query = search_query.lower()
        filtered_df = filtered_df[
            filtered_df["Task Title"].str.lower().str.contains(query) | 
            filtered_df["Details"].str.lower().str.contains(query)
        ]
        
    # 2. Date Range
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_d, end_d = date_range
        filtered_df = filtered_df[(filtered_df["Date"] >= start_d) & (filtered_df["Date"] <= end_d)]
    elif isinstance(date_range, date):
        # Single date selected
        filtered_df = filtered_df[filtered_df["Date"] == date_range]
        
    # 3. User Filter
    if selected_users:
        filtered_df = filtered_df[filtered_df["Executed By"].isin(selected_users)]
        
    # 4. Action Filter
    if selected_actions:
        filtered_df = filtered_df[filtered_df["Action"].isin(selected_actions)]
        
    # Display Results
    st.markdown(f"**Showing {len(filtered_df)} log entries matching filters:**")
    
    if filtered_df.empty:
        st.warning("No audit logs match the current filters. Clear filters to see all logs.")
    else:
        # Format for UI display
        display_df = filtered_df.copy()
        display_df["Timestamp"] = display_df["Timestamp"].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if x else "")
        display_df = display_df[["Log ID", "Action", "Task ID", "Task Title", "Executed By", "Timestamp", "Details"]]
        
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        
        # Export options for the FILTERED view
        st.markdown("<div class='section-header'>Export Filtered View</div>", unsafe_allow_html=True)
        
        exp_col1, exp_col2, exp_col3 = st.columns(3)
        
        with exp_col1:
            csv_data = filtered_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📁 Export Filtered CSV",
                data=csv_data,
                file_name=f"tasktracker_filtered_audit_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
            
        with exp_col2:
            # Styled excel for the filtered logs
            xlsx_buffer = io_export_logs_excel(filtered_df)
            st.download_button(
                label="📈 Export Filtered Excel",
                data=xlsx_buffer,
                file_name=f"tasktracker_filtered_audit_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        with exp_col3:
            # We can generate a simple PDF document for the audit logs
            pdf_data = io_export_logs_pdf(filtered_df)
            st.download_button(
                label="📄 Export Filtered PDF",
                data=pdf_data,
                file_name=f"tasktracker_filtered_audit_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
