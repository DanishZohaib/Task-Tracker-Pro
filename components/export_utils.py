import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from config import FOOTER_CREATOR, APP_NAME, TAGLINE

def export_to_csv(df: pd.DataFrame) -> bytes:
    """Export tasks dataframe to CSV bytes."""
    return df.to_csv(index=False).encode('utf-8')

def export_to_excel(df: pd.DataFrame) -> bytes:
    """Export tasks dataframe to styled Excel bytes."""
    output = io.BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Tasks Audit Trail', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Tasks Audit Trail']
        
        # Styles
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo
        cell_font = Font(name='Segoe UI', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Format headers
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Format content rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                
                # Format Dates/Times
                val = cell.value
                if isinstance(val, (datetime, pd.Timestamp)):
                    cell.value = val.strftime('%Y-%m-%d %H:%M:%S')
                    cell.alignment = center_align
                elif col_idx in [1, 4, 11]:  # ID, Status, Priority, etc.
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    return output.getvalue()

def export_to_pdf(df: pd.DataFrame, kpis: dict) -> bytes:
    """Export tasks and summary KPIs to a PDF using ReportLab."""
    pdf_buffer = io.BytesIO()
    
    # 0.75-inch margins
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=letter,
        rightMargin=54, leftMargin=54, 
        topMargin=54, bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=12,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=15
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1E293B'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#1F2937')
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#94A3B8'),
        alignment=1 # Centered
    )

    story = []
    
    # Header Branding
    story.append(Paragraph(APP_NAME, title_style))
    story.append(Paragraph(TAGLINE, subtitle_style))
    story.append(Spacer(1, 10))
    
    # Executive KPI Dashboard Grid
    kpi_data = [
        [
            Paragraph("<b>Total Tasks:</b>", body_style), Paragraph(str(kpis.get("total_tasks", 0)), body_style),
            Paragraph("<b>Completion Rate:</b>", body_style), Paragraph(f"{kpis.get('completion_rate', 0.0):.1f}%", body_style)
        ],
        [
            Paragraph("<b>Pending Tasks:</b>", body_style), Paragraph(str(kpis.get("pending_tasks", 0)), body_style),
            Paragraph("<b>Completed Tasks:</b>", body_style), Paragraph(str(kpis.get("completed_tasks", 0)), body_style)
        ],
        [
            Paragraph("<b>Overdue Tasks:</b>", body_style), Paragraph(str(kpis.get("overdue_tasks", 0)), body_style),
            Paragraph("<b>Created This Month:</b>", body_style), Paragraph(str(kpis.get("created_this_month", 0)), body_style)
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[1.5*inch, 1.0*inch, 1.8*inch, 1.2*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#F1F5F9')),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(Paragraph("Executive Performance Metrics", section_title_style))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # Task List Section
    story.append(Paragraph("Detailed Tasks Report", section_title_style))
    
    # Column mapping & prep
    # Select key fields to display: ID, Title, Status, Priority, Created By, Date
    # ReportLab table cells need to wrap nicely using Paragraphs
    table_headers = ["ID", "Title", "Status", "Priority", "Created By", "Created Date"]
    table_data = [[Paragraph(h, header_style) for h in table_headers]]
    
    for _, row in df.iterrows():
        # Shorten titles if too long
        title_text = str(row['Task Title'])
        if len(title_text) > 40:
            title_text = title_text[:37] + "..."
            
        created_dt = row['Created DateTime']
        created_str = ""
        if isinstance(created_dt, (datetime, pd.Timestamp)):
            created_str = created_dt.strftime('%Y-%m-%d')
        else:
            created_str = str(created_dt)[:10]
            
        row_cells = [
            Paragraph(str(row['Task ID']), body_style),
            Paragraph(title_text, body_style),
            Paragraph(str(row['Status']), body_style),
            Paragraph(str(row['Priority']), body_style),
            Paragraph(str(row['Created By']), body_style),
            Paragraph(created_str, body_style)
        ]
        table_data.append(row_cells)
        
    # Column widths adding up to 7.0 inches (total printable width is 8.5 - 2*0.75 = 7.0 inches)
    task_table = Table(table_data, colWidths=[0.5*inch, 2.3*inch, 0.9*inch, 0.9*inch, 1.2*inch, 1.2*inch])
    task_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(task_table)
    story.append(Spacer(1, 30))
    
    # Footer and Branding Credit
    story.append(Paragraph(f"{FOOTER_CREATOR} | TaskTracker Pro &copy; 2026", footer_style))
    
    # Build Document
    doc.build(story)
    
    return pdf_buffer.getvalue()
